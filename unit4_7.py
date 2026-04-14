from openpyxl import load_workbook
wb=load_workbook('students.xlsx');ws=wb.active
m=f=0
for r in ws.iter_rows(min_row=2,values_only=True):
 if r[2]=='M':m+=1
 else:f+=1
import matplotlib.pyplot as plt
plt.bar(['M','F'],[m,f]);plt.show()