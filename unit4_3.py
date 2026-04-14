from openpyxl import load_workbook
wb=load_workbook('students.xlsx');ws=wb.active
for row in ws.iter_rows(min_row=2,values_only=True):
 if row[5]>=20:print(row)