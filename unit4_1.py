from openpyxl import load_workbook
wb=load_workbook('students.xlsx');ws=wb.active
for col in ws.iter_cols(1,ws.max_column):print(col[0].value,type(col[1].value))