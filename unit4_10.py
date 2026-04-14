import re
from openpyxl import load_workbook
wb=load_workbook('students.xlsx');ws=wb.active
for r in ws.iter_rows(min_row=2,values_only=True):
 if re.match(r'^91\d{10}$',str(r[4])):print(r)