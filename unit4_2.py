from openpyxl import load_workbook
wb=load_workbook('students.xlsx');ws=wb.active
r=[];m=[];rm=[]
for row in ws.iter_rows(min_row=2,values_only=True):
 if row[6]=='Rajkot':r.append(row)
 if row[2]=='M':m.append(row)
 if row[6]=='Rajkot' and row[2]=='M':rm.append(row)
print(r,m,rm)